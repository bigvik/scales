'''Модуль сохранения данных.
Класс Datasaver имеет методы set_data() и save_data()'''


import sqlite3
import config


class Datasaver:
    '''
    Класс модели сохранения данных
    '''
    def __init__(self, data) -> None:
        if data: self.data = data
        self.db = config.DB

    def set_data(self,data):
        self.data = data

    def save_data(self):
        '''
        Сохраняет данные в xls или sqlite
        в зависимости от настройки DB в файле config.py
        '''
        destinations = {'xls':self.to_xls, 'sql':self.to_sql}
        destinations[self.db]()

    def to_xls(self):
        import os
        if not os.path.isfile('measurements.xlsx'):
            from openpyxl import Workbook
            wb = Workbook()
            wb.save('measurements.xlsx')
        from openpyxl import load_workbook
        xl = load_workbook('measurements.xlsx')
        xl_sheet = xl.active
        xl_sheet.append(self.data)
        xl.save('measurements.xlsx')

    def to_xlsbydate(self, date):
        print(date)
        import os
        if not os.path.isfile(f'xlsx/measurements_{date}.xlsx'):
            from openpyxl import Workbook
            wb = Workbook()
            wb.save(f'xlsx/measurements_{date}.xlsx')
        from openpyxl import load_workbook
        con = sqlite3.connect('measurements.db')
        cur = con.cursor()
        cur.execute(
            f"""SELECT date,brutto,netto,tara,dest FROM "weight" WHERE "date" LIKE "%{date}%" """
        )
        rows = cur.fetchall()
        print(len(rows))
        xl = load_workbook(f'xlsx/measurements_{date}.xlsx')
        xl_sheet = xl.active
        for row in rows:
            xl_sheet.append(row)
        xl.save(f'xlsx/measurements_{date}.xlsx')

    def to_sql(self):
        con = sqlite3.connect('measurements.db')
        cur = con.cursor()
        cur.execute("""CREATE TABLE IF NOT EXISTS weight(
                    id INTEGER PRIMARY KEY,
                    date TEXT,
                    brutto INT,
                    netto INT,
                    tara INT,
                    dest TEXT);
                    """)
        con.commit()
        cur.execute("INSERT INTO weight(date, brutto, netto, tara, dest) VALUES(?, ?, ?, ?, ?);", self.data)
        con.commit()


if __name__ == "__main__":
    pass